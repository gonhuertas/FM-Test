"""Parse BRH quarterly financial ratio sheets (sysratfinclé series).

Reads all brh_trim*.xlsx/.xls files in data/raw/, extracts key prudential and
profitability ratios from every sysratfinclé sheet, deduplicates across files,
and saves a tidy CSV to data/processed/brh_ratios.csv.

Key design decisions:
- Column layout is detected dynamically from row 6 of each sheet, so the parser
  handles the bank entries/exits that occurred over 2000-2025 without hardcoding.
- Both .xlsx (openpyxl) and .xls (xlrd) formats are supported.
- Row positions for all 19 metrics are stable across all years (verified).

Usage:
    python scripts/parse_brh_ratios.py
"""

import io
import re
from pathlib import Path

import openpyxl
import xlrd
import pandas as pd

# ── Paths ─────────────────────────────────────────────────────────────────────

ROOT        = Path(__file__).parent.parent
RAW_DIR     = ROOT / "data" / "raw"
OUTPUT_FILE = ROOT / "data" / "processed" / "brh_ratios.csv"

# ── Bank name normalisation ───────────────────────────────────────────────────
# Some banks were renamed or relabelled across vintages of the report.

BANK_ALIASES: dict[str, str] = {
    "SGBEL":       "SOGEBL",    # early abbreviation for Société Générale de Banque Locale
    "SOCABL":      "SOGEBL",    # another early variant (pre-2005)
    "TOTAL":       "SYSTÈME",   # pre-2008 reports called the system aggregate "TOTAL"
    "BHD *":       "BHD",       # footnoted variant
    "PROMOBK **":  "PROMOBK",   # footnoted variant
}

# These column labels are structural subtotals, not individual banks — skip them.
SKIP_LABELS = {"SOUS-TOTAL"}

# ── French month abbreviations → month number ─────────────────────────────────

MONTH_MAP: dict[str, int] = {
    "jan": 1,  "janv": 1,
    "fév": 2,  "fev": 2,  "fevr": 2,
    "mars": 3, "mar": 3,
    "avr": 4,  "avril": 4,
    "mai": 5,
    "juin": 6, "jun": 6,
    "juil": 7, "juill": 7, "juillet": 7,
    "août": 8, "aout": 8,
    "sept": 9, "sep": 9,
    "oct": 10, "octo": 10,
    "nov": 11,
    "déc": 12, "dec": 12, "dèc": 12,
}

# ── Metric definitions ────────────────────────────────────────────────────────
# Row numbers are 1-indexed and stable across ALL reports (2000–2025, verified).
#
# period_type:
#   "year_end" — balance-sheet / stock metric at the fiscal year-end date
#   "quarter"  — flow metric for the current quarter (TRIM current)
#   "cumul"    — flow metric cumulated over the full fiscal year (CUMUL current)

METRICS: list[dict] = [
    # ── STRUCTURE FINANCIÈRE ──────────────────────────────────────────────────
    {"row": 10, "name": "equity_to_assets",    "label": "Capital Ratio (Equity / Assets)",        "period_type": "year_end", "unit": "ratio"},
    {"row": 14, "name": "deposits_to_assets",  "label": "Deposits / Total Assets",                "period_type": "year_end", "unit": "ratio"},
    # ── QUALITÉ DE L'ACTIF ────────────────────────────────────────────────────
    {"row": 20, "name": "npl_ratio_gross",     "label": "Gross NPL Ratio",                        "period_type": "year_end", "unit": "ratio"},
    {"row": 25, "name": "provision_coverage",  "label": "Provision Coverage (provisions / NPLs)", "period_type": "year_end", "unit": "ratio"},
    {"row": 30, "name": "net_npl_to_equity",   "label": "Net NPL / Equity",                       "period_type": "year_end", "unit": "ratio"},
    # ── RENTABILITÉ — quarterly ───────────────────────────────────────────────
    {"row": 35, "name": "roa_q",               "label": "ROA (quarterly, annualised)",            "period_type": "quarter",  "unit": "ratio"},
    {"row": 41, "name": "roe_q",               "label": "ROE (quarterly, annualised)",            "period_type": "quarter",  "unit": "ratio"},
    {"row": 48, "name": "nim_q",               "label": "Net Interest Margin (quarterly)",        "period_type": "quarter",  "unit": "ratio"},
    {"row": 54, "name": "avg_loan_yield_q",    "label": "Average Loan Yield (quarterly)",         "period_type": "quarter",  "unit": "ratio"},
    {"row": 66, "name": "avg_deposit_rate_q",  "label": "Average Deposit Rate (quarterly)",       "period_type": "quarter",  "unit": "ratio"},
    {"row": 74, "name": "cost_to_income_q",    "label": "Cost-to-Income (quarterly)",             "period_type": "quarter",  "unit": "ratio"},
    {"row": 80, "name": "productivity_q",      "label": "Employee Productivity / quarter (HTG '000)", "period_type": "quarter", "unit": "HTG_thousands"},
    # ── RENTABILITÉ — cumulative ──────────────────────────────────────────────
    {"row": 37, "name": "roa_cumul",           "label": "ROA (fiscal year cumulative)",           "period_type": "cumul",    "unit": "ratio"},
    {"row": 43, "name": "roe_cumul",           "label": "ROE (fiscal year cumulative)",           "period_type": "cumul",    "unit": "ratio"},
    {"row": 50, "name": "nim_cumul",           "label": "Net Interest Margin (cumulative)",       "period_type": "cumul",    "unit": "ratio"},
    {"row": 56, "name": "avg_loan_yield_c",    "label": "Average Loan Yield (cumulative)",        "period_type": "cumul",    "unit": "ratio"},
    {"row": 68, "name": "avg_deposit_rate_c",  "label": "Average Deposit Rate (cumulative)",      "period_type": "cumul",    "unit": "ratio"},
    {"row": 76, "name": "cost_to_income_c",    "label": "Cost-to-Income (cumulative)",            "period_type": "cumul",    "unit": "ratio"},
    {"row": 82, "name": "productivity_c",      "label": "Employee Productivity / year (HTG '000)", "period_type": "cumul",  "unit": "HTG_thousands"},
]

# Fragments that must appear in specific rows as a structural sanity check
EXPECTED_LABELS: dict[int, str] = {
    7:  "STRUCTURE",
    17: "QUALIT",       # covers both "QUALITÉ" and any encoding variant
    33: "RENTABILIT",   # covers both "RENTABILITÉ" and variants
}


# ── Low-level cell readers ────────────────────────────────────────────────────

def _open_xlsx(path: Path) -> openpyxl.Workbook:
    """Open an xlsx workbook via BytesIO to avoid openpyxl's extension-based checks.

    Some BRH files are valid xlsx but carry a .xls extension, which causes
    openpyxl to raise InvalidFileException when given the path directly.
    """
    return openpyxl.load_workbook(
        io.BytesIO(path.read_bytes()), read_only=True, data_only=True
    )


def _cells_from_xlsx(wb: openpyxl.Workbook, sheet_name: str) -> dict[tuple, object]:
    ws = wb[sheet_name]
    return {
        (cell.row, cell.column): cell.value
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    }


def _cells_from_xls(wb: xlrd.Book, sheet_name: str) -> dict[tuple, object]:
    ws = wb.sheet_by_name(sheet_name)
    cells = {}
    for r in range(ws.nrows):
        for c in range(ws.ncols):
            val = ws.cell_value(r, c)
            if val != "" and val is not None:
                cells[(r + 1, c + 1)] = val   # convert to 1-indexed to match openpyxl
    return cells


# ── Parsing helpers ───────────────────────────────────────────────────────────

def parse_sheet_date(sheet_name: str) -> pd.Timestamp | None:
    """Extract the quarter-end date from a sheet name.

    Handles all format variants found across 25 years of reports:
      'sysratfinclé sept 25'    space-separated, 2-digit year
      'sysratfinclé déc. 24'    month with trailing period
      'sysratfinclé.juin. 17'   dots as separators
      'sysratfinclédéc. 14'     no space between prefix and month
      'sysratfinclésept.00 '    2-digit year with trailing space (year 2000)
    """
    name = sheet_name.lower().strip()
    for month_str, month_num in MONTH_MAP.items():
        # Allow optional dot after month, then optional whitespace, then 2-digit year
        m = re.search(re.escape(month_str) + r"\.?\s*(\d{2})\b", name)
        if m:
            yy = int(m.group(1))
            # Two-digit years: 00-49 → 2000-2049, 50-99 → 1950-1999
            # The BRH archive starts at fiscal year 2000 (sheets go back to late 1999)
            year = (1900 + yy) if yy >= 50 else (2000 + yy)
            return pd.Timestamp(year=year, month=month_num, day=1) + pd.offsets.MonthEnd(0)
    return None


def detect_bank_columns(cells: dict) -> dict[int, str]:
    """Read row 6 and return {col_1idx: canonical_bank_name}.

    Applies BANK_ALIASES (e.g. TOTAL->SYSTÈME, SGBEL->SOGEBL) and drops
    SOUS-TOTAL subtotal columns.  Works regardless of how many banks are in
    the report (column count varied from 14 to 18 over 2000-2025).
    """
    col_map = {}
    for (r, c), val in cells.items():
        if r != 6:
            continue
        if not isinstance(val, str):
            continue
        name = val.strip()
        if not name:
            continue
        name = BANK_ALIASES.get(name, name)   # normalise aliases
        if name not in SKIP_LABELS:
            col_map[c] = name
    return col_map


def safe_float(value) -> float | None:
    """Convert a cell value to float; return None for blanks / N/A."""
    if value is None:
        return None
    if isinstance(value, str) and value.strip() in ("", " ", "N/A", "N/D"):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def validate_structure(cells: dict, sheet_name: str) -> bool:
    """Return True if the expected section headers are in the right rows."""
    for row, fragment in EXPECTED_LABELS.items():
        val = str(cells.get((row, 1), ""))
        if fragment.lower() not in val.lower():
            return False
    return True


def parse_sheet(cells: dict, date: pd.Timestamp, bank_cols: dict[int, str]) -> list[dict]:
    """Extract all metric × bank combinations from a pre-loaded cell dict."""
    records = []
    for metric in METRICS:
        r = metric["row"]
        for col, bank in bank_cols.items():
            value = safe_float(cells.get((r, col)))
            # Zero is not a valid ratio — treat as missing (inactive bank or N/A)
            if metric["unit"] == "ratio" and value == 0.0:
                value = None
            records.append({
                "date":        date,
                "metric":      metric["name"],
                "label":       metric["label"],
                "period_type": metric["period_type"],
                "unit":        metric["unit"],
                "bank":        bank,
                "value":       value,
            })
    return records


# ── File-level parsing ────────────────────────────────────────────────────────

def _is_xlsx_format(path: Path) -> bool:
    """Detect actual file format by magic bytes (not file extension).

    xlsx/xlsm are zip archives starting with PK (0x50 0x4B).
    Legacy .xls (BIFF) files start with 0xD0 0xCF.
    Some BRH files have a .xls extension but are actually xlsx.
    """
    with open(path, "rb") as f:
        magic = f.read(2)
    return magic == b"PK"


def parse_file(path: Path) -> tuple[list[dict], int, int]:
    """Parse all sysratfinclé sheets in one report file.

    Returns (records, n_parsed, n_skipped).
    """
    records   = []
    n_parsed  = 0
    n_skipped = 0
    use_xlsx  = _is_xlsx_format(path)

    # Always use BytesIO for xlsx to avoid openpyxl's extension-based validation.
    # Some BRH files are valid xlsx but carry a .xls extension, which trips openpyxl.
    try:
        if use_xlsx:
            wb          = _open_xlsx(path)
            sheet_names = [s for s in wb.sheetnames if "sysratfincl" in s.lower()]
            get_cells   = lambda name: _cells_from_xlsx(wb, name)
            close       = wb.close
        else:
            wb          = xlrd.open_workbook(str(path))
            sheet_names = [s for s in wb.sheet_names() if "sysratfincl" in s.lower()]
            get_cells   = lambda name: _cells_from_xls(wb, name)
            close       = lambda: None
    except Exception as e:
        print(f"    ERROR: could not open {path.name}: {e}")
        return [], 0, 0

    for sheet_name in sheet_names:
        date = parse_sheet_date(sheet_name)
        if date is None:
            n_skipped += 1
            continue

        cells = get_cells(sheet_name)

        if not validate_structure(cells, sheet_name):
            n_skipped += 1
            continue

        bank_cols = detect_bank_columns(cells)
        records.extend(parse_sheet(cells, date, bank_cols))
        n_parsed += 1

    close()
    return records, n_parsed, n_skipped


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    raw_files = sorted(RAW_DIR.glob("brh_trim*.xl*"))
    print(f"Found {len(raw_files)} report files in data/raw/\n")

    all_records: list[dict] = []
    total_parsed  = 0
    total_skipped = 0

    for fpath in raw_files:
        records, n_parsed, n_skipped = parse_file(fpath)
        all_records.extend(records)
        total_parsed  += n_parsed
        total_skipped += n_skipped
        print(f"  {fpath.name:<30} {n_parsed} sheets parsed, {n_skipped} skipped")

    print(f"\nTotal: {total_parsed} sheets parsed, {total_skipped} skipped")

    df = pd.DataFrame(all_records)

    # Deduplicate: the same quarter appears in multiple files (each file contains
    # up to 8 quarters of history).  Values are identical across files, so just
    # keep one row per (date, metric, bank).
    before = len(df)
    df = (
        df
        .dropna(subset=["value"])
        .sort_values(["metric", "bank", "date"])
        .drop_duplicates(subset=["date", "metric", "bank"], keep="last")
        .reset_index(drop=True)
    )
    print(f"Rows after dedup: {len(df):,}  (removed {before - len(df):,} duplicates)")

    df.to_csv(OUTPUT_FILE, index=False)

    print(f"\nSaved {len(df):,} rows -> {OUTPUT_FILE}")
    print(f"Date range : {df['date'].min().date()} -- {df['date'].max().date()}")
    print(f"Metrics    : {df['metric'].nunique()}")
    print(f"Banks      : {sorted(df['bank'].unique())}")


if __name__ == "__main__":
    main()
