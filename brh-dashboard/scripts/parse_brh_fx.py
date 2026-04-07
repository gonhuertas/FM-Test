"""Parse BRH monthly FX net open position data (posinette sheets).

Reads all brh_trim*.xlsx/.xls files in data/raw/, extracts monthly FX position
and breach counts for every bank, deduplicates across overlapping files, and
saves to data/processed/brh_fx_positions.csv.

Layout of the posinette sheet (stable across all years):
  - Row N:    "Trim X YYYY" at col 3 (current quarter) and col 11/12 (prior quarter)
  - Row N+1:  Month names  — current quarter at cols 3, 5, 7; prior at q_prev_col, +2, +4
  - Row N+3:  Bank data starts (col 2 = bank name; col+0 = position, col+1 = days exceeded)
  - Positions are expressed as a ratio of equity (e.g. 0.04 = 4%)
  - Days exceeded = number of calendar days in the month the limit was breached

The regulatory limit (Circulaire 81-3, later amended by 81-4 and 81-6) is not
embedded as a number in the sheet; days_exceeded > 0 is used as the breach flag.

Date handling:
  - Oct/Nov/Dec are fiscal-year Q1; calendar year = fiscal_year_in_header − 1
  - All other months: calendar year = fiscal_year_in_header

Usage:
    python scripts/parse_brh_fx.py
"""

import io
import re
from pathlib import Path

import openpyxl
import xlrd
import pandas as pd

# ── Paths ─────────────────────────────────────────────────────────────────────

ROOT        = Path(__file__).parent.parent
RAW_DIR     = ROOT / "data" / "raw"
OUTPUT_FILE = ROOT / "data" / "processed" / "brh_fx_positions.csv"

# ── Constants ─────────────────────────────────────────────────────────────────

BANK_ALIASES: dict[str, str] = {
    "SGBEL":  "SOGEBL",
    "SOCABL": "SOGEBL",
}

# These are system aggregates, not individual banks — skip them.
SKIP_LABELS = {"SOUS-TOTAL", "SYSTÈME", "TOTAL"}

MONTH_MAP: dict[str, int] = {
    "janvier": 1, "janv": 1, "jan": 1,
    "février": 2, "fevrier": 2, "fev": 2, "fevr": 2,
    "mars": 3,
    "avril": 4, "avr": 4,
    "mai": 5,
    "juin": 6, "jun": 6,
    "juillet": 7, "juill": 7, "juil": 7,
    "août": 8, "aout": 8,
    "septembre": 9, "sept": 9, "sep": 9,
    "octobre": 10, "oct": 10,
    "novembre": 11, "nov": 11,
    "décembre": 12, "decembre": 12, "dec": 12,
}


# ── Low-level file helpers ────────────────────────────────────────────────────

def _is_xlsx_format(path: Path) -> bool:
    with open(path, "rb") as f:
        return f.read(2) == b"PK"


def _open_xlsx(path: Path) -> openpyxl.Workbook:
    """Always load via BytesIO to bypass openpyxl's extension-based validation."""
    return openpyxl.load_workbook(
        io.BytesIO(path.read_bytes()), read_only=True, data_only=True
    )


def _cells_from_xlsx(wb: openpyxl.Workbook, sheet_name: str) -> dict[tuple, object]:
    ws = wb[sheet_name]
    return {
        (cell.row, cell.column): cell.value
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    }


def _cells_from_xls(wb: xlrd.Book, sheet_name: str) -> dict[tuple, object]:
    ws = wb.sheet_by_name(sheet_name)
    cells = {}
    for r in range(ws.nrows):
        for c in range(ws.ncols):
            val = ws.cell_value(r, c)
            if val != "" and val is not None:
                cells[(r + 1, c + 1)] = val   # 1-indexed to match openpyxl
    return cells


# ── Parsing helpers ───────────────────────────────────────────────────────────

def safe_float(value) -> float | None:
    """Convert a cell value to float, handling blanks and string-encoded numbers."""
    if value is None:
        return None
    if isinstance(value, str) and value.strip() in ("", " ", "N/A", "N/D"):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def parse_month_num(text) -> int | None:
    """Extract a month number from a French month header cell (handles footnote digits)."""
    # Strip digits and punctuation to isolate the word
    t = re.sub(r'[\d\.\s]', '', str(text).lower())
    # Try longest token first to avoid "mar" matching "mars" then "avr" not matching "avril"
    for month_str, month_num in sorted(MONTH_MAP.items(), key=lambda x: -len(x[0])):
        if month_str in t:
            return month_num
    return None


def parse_trim_year(text) -> tuple[int, int] | None:
    """Parse 'Trim X YYYY' (or 'Trim X YY', or 'Trim X, YYYY') -> (quarter, fiscal_year)."""
    s = str(text).lower()
    # Allow optional comma/space between quarter digit and year
    m = re.search(r'trim\s*(\d)[,\s]+(\d{4})', s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r'trim\s*(\d)[,\s]+(\d{2})\b', s)
    if m:
        yy = int(m.group(2))
        year = 1900 + yy if yy >= 50 else 2000 + yy
        return int(m.group(1)), year
    return None


def make_date(fiscal_year: int, month_num: int) -> pd.Timestamp:
    """Convert (fiscal_year, month_num) to a calendar month-end Timestamp.

    Haiti's fiscal year runs October–September.  Oct/Nov/Dec are Q1, so their
    calendar year is fiscal_year − 1.  All other months match the fiscal year.
    """
    cal_year = fiscal_year - 1 if month_num >= 10 else fiscal_year
    return pd.Timestamp(year=cal_year, month=month_num, day=1) + pd.offsets.MonthEnd(0)


def _is_bank_name(val) -> bool:
    """Return True if val looks like a bank abbreviation, not a footnote or blank."""
    if not isinstance(val, str):
        return False
    v = val.strip()
    return bool(v) and v[0].isupper() and not v[0].isdigit() and len(v) <= 15


# ── Sheet-level parser ────────────────────────────────────────────────────────

def parse_posinette(cells: dict) -> list[dict]:
    """Extract monthly records from a pre-loaded posinette cell dict.

    Returns a list of dicts with keys: date, bank, fx_position, days_exceeded.
    """
    # ── Locate the Trim header row ────────────────────────────────────────────
    # This is the row where col 3 says "Trim X YYYY".
    header_row = None
    for r in range(1, 65):
        val = str(cells.get((r, 3), "")).lower()
        if "trim" in val and re.search(r'\d{2}', val):
            header_row = r
            break
    if header_row is None:
        return []

    # ── Parse the two quarter headers ────────────────────────────────────────
    q_curr_info = parse_trim_year(cells.get((header_row, 3), ""))
    if q_curr_info is None:
        return []
    _q_curr, q_curr_fy = q_curr_info

    # Prior quarter starts somewhere to the right (col 9–22).
    q_prev_start_col = None
    q_prev_fy = None
    for c in range(9, 23):
        val = cells.get((header_row, c), "")
        if "trim" in str(val).lower() and re.search(r'\d{2}', str(val)):
            info = parse_trim_year(val)
            if info:
                _, q_prev_fy = info
                q_prev_start_col = c
            break

    month_row  = header_row + 1
    data_start = header_row + 3   # one blank row between month headers and bank data

    # ── Determine month columns ───────────────────────────────────────────────
    # Current quarter: always starts at col 3 (position), col 4 (days), col 5, 6, col 7, 8
    q_curr_cols: list[tuple[int, int]] = []   # (pos_col, month_num)
    for col in [3, 5, 7]:
        m_num = parse_month_num(cells.get((month_row, col), ""))
        if m_num:
            q_curr_cols.append((col, m_num))

    # Prior quarter: starts at q_prev_start_col, then +2, +4
    q_prev_cols: list[tuple[int, int]] = []
    if q_prev_start_col is not None and q_prev_fy is not None:
        for offset in [0, 2, 4]:
            col = q_prev_start_col + offset
            m_num = parse_month_num(cells.get((month_row, col), ""))
            if m_num:
                q_prev_cols.append((col, m_num))

    if not q_curr_cols:
        return []

    # ── Extract bank-level data ───────────────────────────────────────────────
    records = []
    for r in range(data_start, data_start + 22):   # at most ~15 banks
        bank_val = cells.get((r, 2))
        if not _is_bank_name(bank_val):
            if bank_val is None:
                continue       # tolerate a blank row
            break              # non-bank text (footnote) signals end of data

        bank = str(bank_val).strip()
        bank = BANK_ALIASES.get(bank, bank)
        if bank in SKIP_LABELS:
            continue

        def _record(col: int, m_num: int, fy: int) -> dict:
            pos  = safe_float(cells.get((r, col)))
            days_raw = safe_float(cells.get((r, col + 1)))
            days = int(days_raw) if days_raw is not None else 0
            # Zero position with no breach is real data (flat position); keep it.
            return {
                "date":         make_date(fy, m_num),
                "bank":         bank,
                "fx_position":  pos,           # ratio to equity, e.g. 0.04 = 4%
                "days_exceeded": days,          # number of days limit was breached
            }

        for col, m_num in q_curr_cols:
            records.append(_record(col, m_num, q_curr_fy))
        for col, m_num in q_prev_cols:
            records.append(_record(col, m_num, q_prev_fy))

    return records


# ── File-level parsing ────────────────────────────────────────────────────────

def parse_file(path: Path) -> tuple[list[dict], str]:
    """Open a report file, parse its posinette sheet, and return (records, status)."""
    use_xlsx = _is_xlsx_format(path)
    try:
        if use_xlsx:
            wb    = _open_xlsx(path)
            names = [s for s in wb.sheetnames if "posinette" in s.lower()]
            if not names:
                wb.close()
                return [], "no posinette sheet"
            cells = _cells_from_xlsx(wb, names[0])
            wb.close()
        else:
            wb    = xlrd.open_workbook(str(path))
            names = [s for s in wb.sheet_names() if "posinette" in s.lower()]
            if not names:
                return [], "no posinette sheet"
            cells = _cells_from_xls(wb, names[0])
    except Exception as e:
        return [], f"ERROR: {e}"

    records = parse_posinette(cells)
    return records, f"{len(records)} records"


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    raw_files = sorted(RAW_DIR.glob("brh_trim*.xl*"))
    print(f"Found {len(raw_files)} report files in data/raw/\n")

    all_records: list[dict] = []

    for fpath in raw_files:
        records, status = parse_file(fpath)
        all_records.extend(records)
        print(f"  {fpath.name:<30} {status}")

    df = pd.DataFrame(all_records)
    if df.empty:
        print("\nNo records parsed.")
        return

    before = len(df)
    df = (
        df
        .sort_values(["bank", "date"])
        .drop_duplicates(subset=["date", "bank"], keep="last")
        .reset_index(drop=True)
    )
    print(f"\nRows after dedup: {len(df):,}  (removed {before - len(df):,} duplicates)")

    df.to_csv(OUTPUT_FILE, index=False)

    print(f"\nSaved {len(df):,} rows -> {OUTPUT_FILE}")
    print(f"Date range  : {df['date'].min().date()} -- {df['date'].max().date()}")
    print(f"Banks       : {sorted(df['bank'].unique())}")
    breaches = (df["days_exceeded"] > 0).sum()
    print(f"Month-bank observations with any breach: {breaches:,} "
          f"({100 * breaches / len(df):.1f}%)")


if __name__ == "__main__":
    main()
