"""
Search X/Twitter and news outlets for discussions on a topic using Grok's built-in tools.
Uses the xai-sdk with x_search (X/Twitter) and web_search (news) tools.
Results are saved to a structured Excel file with deduplication.
"""

from datetime import datetime
from pathlib import Path

import openpyxl
from pydantic import BaseModel

from xai_sdk.sync.client import Client
from xai_sdk.chat import user
from xai_sdk.tools import x_search, web_search  # built-in tools for searching X and the web

from credentials import grok_token


VALID_TAGS = ["Violence", "Complaints", "Protest", "Government", "Economy", "Media", "Disruptions", "Misc"]

TAG_DEFINITIONS = """
Assign exactly one tag per quote from the following list:
- Violence: specific recent violent events (named gangs, named locations, specific incidents)
- Complaints: general grievances, frustration, condemnation without specific incident detail
- Protest: existing protests OR explicit calls to protest/march/mobilize/boycott
- Government: official actions, decrees, advisory bodies, PM decisions
- Economy: prices, wages, inflation, supply, black market, cost of living
- Media: factual neutral reporting of developments
- Disruptions: transport chaos, station lines, operational breakdown
- Misc: genuinely cross-cutting or unclear
""".strip()


class Quote(BaseModel):
    text: str   # the quote text, including handle/attribution
    tag: str    # one of the VALID_TAGS above


class NewsQuote(BaseModel):
    text: str   # the quote text, in format "Outlet: 'quote text'"
    tag: str    # one of the VALID_TAGS above


class XSearchResult(BaseModel):
    summary: str          # main themes and sentiments
    highlights: str       # key points being made
    quotes: list[Quote]   # direct quotes with handles, each tagged
    consensus: str        # agreements or disagreements


class WebSearchResult(BaseModel):
    summary: str               # main themes from news coverage
    highlights: str            # key facts, numbers, or events reported
    sources: str               # cited outlet articles, in format "Outlet — Headline", one per line
    news_quotes: list[NewsQuote]  # notable quotes from articles, each tagged
    consensus: str             # what news coverage broadly agrees on or disputes

# ── Configuration ─────────────────────────────────────────────────────────────

TOPIC = "haiti fuel prices"  # ← change this

FROM_DATE = datetime(2026, 4, 2)
FROM_DATE_NEWS = datetime(2026, 4, 2)
TO_DATE   = datetime(2026, 4, 3)

MODEL      = "grok-4-1-fast"
MULTI_PASS = 0  # 0 = single search (testing); 1 = multiple passes + synthesis
N_RUNS     = 2  # number of passes when MULTI_PASS = 1

PRIOR_RUNS = 3  # how many past runs to load as context for Grok

EXCEL_PATH = Path("output/x_search_log.xlsx")

# Domains to search for news coverage
NEWS_DOMAINS = [
    "lenouvelliste.com",
    "haitilibre.com",
    "rezonodwes.com",
    "icihaiti.com",
    "ayibopost.com",
]

# ── Load prior context from Excel ─────────────────────────────────────────────

def load_prior_context(excel_path: Path, n: int) -> tuple[str, str]:
    """
    Read the last n rows from the Runs and News Runs sheets and format them
    as context strings to prepend to the Grok prompts.
    Returns (x_context, news_context); either may be "" if no history exists.
    """
    if not excel_path.exists():
        return "", ""

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    def read_last(sheet_name: str) -> list[dict]:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        # columns: Timestamp, Topic, From Date, To Date, Model, Summary, Highlights, Consensus
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]
        rows = rows[-n:]
        return [
            {"timestamp": r[0], "from_date": r[2], "to_date": r[3],
             "summary": r[5], "highlights": r[6], "consensus": r[7]}
            for r in rows
        ]

    def fmt(rows: list[dict]) -> str:
        parts = []
        for r in rows:
            parts.append(
                f"[{r['timestamp']} | {r['from_date']} to {r['to_date']}]\n"
                f"Summary: {r['summary']}\n"
                f"Highlights: {r['highlights']}\n"
                f"Consensus: {r['consensus']}"
            )
        return "\n\n".join(parts)

    x_ctx   = fmt(read_last("Runs"))
    news_ctx = fmt(read_last("News Runs"))
    wb.close()
    return x_ctx, news_ctx


x_prior, news_prior = load_prior_context(EXCEL_PATH, PRIOR_RUNS)

# Preamble injected into prompts when prior history exists
def context_block(prior: str, n: int, source: str) -> str:
    if not prior:
        return ""
    return (
        f"PRIOR RESEARCH CONTEXT — last {n} {source} runs already logged.\n"
        f"Do not characterize anything as a 'first', 'unprecedented', or 'new development' "
        f"if it appears in this record. Only flag something as new if it genuinely does not "
        f"appear below.\n\n"
        f"{prior}\n\n"
        f"---\n\n"
    )

x_context_block    = context_block(x_prior,    PRIOR_RUNS, "X/Twitter")
news_context_block = context_block(news_prior, PRIOR_RUNS, "news")

# ── X/Twitter Search ───────────────────────────────────────────────────────────

client = Client(api_key=grok_token)

x_prompt = (
    x_context_block
    + f"""Search X/Twitter for recent discussions about: "{TOPIC}".
    Provide:
    - summary: main themes and sentiments in the discussion
    - highlights: key points and notable arguments being made
    - quotes: direct quotes from notable or representative posts, including handles; each quote must have a tag
    - consensus: emerging agreements or major points of disagreement

    Be thorough and comprehensive. Return as much detail as you find relevant. Return at least 8 (distinct) quotes if possible. Focus especially on the potential for social unrest, uprisings, or protests.
    Translate any non-English content into English. For quotes, append the original text in parentheses after the translation.

    """
    + TAG_DEFINITIONS
)

if MULTI_PASS == 0:
    print("Single X search pass...")
    chat = client.chat.create(
        model=MODEL,
        tools=[x_search(from_date=FROM_DATE, to_date=TO_DATE)],
        max_turns=10,
    )
    chat.append(user(x_prompt))
    _, x_final = chat.parse(XSearchResult)

else:
    x_raw_runs: list[XSearchResult] = []

    for i in range(N_RUNS):
        print(f"X search pass {i + 1}/{N_RUNS}...")
        chat = client.chat.create(
            model=MODEL,
            tools=[x_search(from_date=FROM_DATE, to_date=TO_DATE)],
            max_turns=10,
        )
        chat.append(user(x_prompt))
        _, parsed = chat.parse(XSearchResult)
        x_raw_runs.append(parsed)

    x_runs_text = "\n\n---\n\n".join(
        f"Pass {i + 1}:\n"
        f"Summary: {r.summary}\n"
        f"Highlights: {r.highlights}\n"
        f"Quotes:\n" + "\n".join(f"  [{q.tag}] {q.text}" for q in r.quotes) + "\n"
        f"Consensus: {r.consensus}"
        for i, r in enumerate(x_raw_runs)
    )

    print("Synthesizing X results...")
    synthesis_chat = client.chat.create(model=MODEL)
    synthesis_chat.append(user(
        f"""You ran {N_RUNS} independent X/Twitter searches about "{TOPIC}".
        Here are the results from each pass:

        {x_runs_text}

        Synthesize these into a single, deduplicated output. Remove repeated quotes or points
        that appear in more than one pass. Keep unique findings from each pass.
        """
    ))
    _, x_final = synthesis_chat.parse(XSearchResult)

# ── News / Web Search ──────────────────────────────────────────────────────────

news_prompt = (
    news_context_block
    + f"""Search news websites for recent coverage about: "{TOPIC}".
    Focus on Haiti-specific journalism from {FROM_DATE_NEWS.strftime('%Y-%m-%d')} to {TO_DATE.strftime('%Y-%m-%d')}.
    Provide:
    - summary: main themes and findings from news reporting
    - highlights: key facts, numbers, or events reported by journalists
    - sources: specific articles cited, in the format "Outlet — Headline", one per line
    - news_quotes: notable direct quotes from journalists or officials cited in the articles, in the format "Outlet: 'quote text'". Prioritize quotes that reference protests, unrest, public reaction, or commentary on government fuel policy. At least 5 quotes if possible. Each quote must have a tag.
    - consensus: what the news coverage broadly agrees on or disputes

    Be thorough. Cite specific outlet names and headlines where possible.
    Translate any non-English content into English. Return only the English translation — do not include the original language anywhere.

    """
    + TAG_DEFINITIONS
)

if MULTI_PASS == 0:
    print("Single news search pass...")
    chat = client.chat.create(
        model=MODEL,
        tools=[web_search(allowed_domains=NEWS_DOMAINS)],
        max_turns=10,
    )
    chat.append(user(news_prompt))
    _, news_final = chat.parse(WebSearchResult)

else:
    news_raw_runs: list[WebSearchResult] = []

    for i in range(N_RUNS):
        print(f"News search pass {i + 1}/{N_RUNS}...")
        chat = client.chat.create(
            model=MODEL,
            tools=[web_search(allowed_domains=NEWS_DOMAINS)],
            max_turns=10,
        )
        chat.append(user(news_prompt))
        _, parsed = chat.parse(WebSearchResult)
        news_raw_runs.append(parsed)

    news_runs_text = "\n\n---\n\n".join(
        f"Pass {i + 1}:\n"
        f"Summary: {r.summary}\n"
        f"Highlights: {r.highlights}\n"
        f"Sources: {r.sources}\n"
        f"News Quotes:\n" + "\n".join(f"  [{q.tag}] {q.text}" for q in r.news_quotes) + "\n"
        f"Consensus: {r.consensus}"
        for i, r in enumerate(news_raw_runs)
    )

    print("Synthesizing news results...")
    news_synthesis_chat = client.chat.create(model=MODEL)
    news_synthesis_chat.append(user(
        f"""You ran {N_RUNS} independent news searches about "{TOPIC}".
        Here are the results from each pass:

        {news_runs_text}

        Synthesize these into a single, deduplicated output. Remove repeated source citations
        that appear in more than one pass. Keep unique findings from each pass.
        """
    ))
    _, news_final = news_synthesis_chat.parse(WebSearchResult)

# ── Print results ──────────────────────────────────────────────────────────────

print("\n=== X/Twitter ===")
print(x_final.summary)
print(x_final.highlights)
for q in x_final.quotes:
    print(f"  [{q.tag}] {q.text}")
print(x_final.consensus)

print("\n=== News ===")
print(news_final.summary)
print(news_final.highlights)
print(news_final.sources)
for q in news_final.news_quotes:
    print(f"  [{q.tag}] {q.text}")
print(news_final.consensus)

# ── Save to Excel ──────────────────────────────────────────────────────────────

EXCEL_PATH.parent.mkdir(exist_ok=True)

timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_or_create_sheet(wb: openpyxl.Workbook, name: str, headers: list[str]):
    """Return existing sheet or create it with a header row."""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    return ws


if EXCEL_PATH.exists():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws_runs         = wb["Runs"]
    ws_quotes       = wb["Quotes"]
    ws_news_runs    = get_or_create_sheet(wb, "News Runs",
                          ["Timestamp", "Topic", "From Date", "To Date", "Model",
                           "Summary", "Highlights", "Consensus"])
    ws_news_sources = get_or_create_sheet(wb, "News Sources",
                          ["Timestamp", "Topic", "Source"])
    ws_news_quotes  = get_or_create_sheet(wb, "News Quotes",
                          ["Timestamp", "Topic", "Quote", "Tag"])
else:
    wb = openpyxl.Workbook()

    ws_runs = wb.active
    ws_runs.title = "Runs"
    ws_runs.append(["Timestamp", "Topic", "From Date", "To Date", "Model",
                    "Summary", "Highlights", "Consensus"])

    ws_quotes = wb.create_sheet("Quotes")
    ws_quotes.append(["Timestamp", "Topic", "Quote", "Tag"])

    ws_news_runs = wb.create_sheet("News Runs")
    ws_news_runs.append(["Timestamp", "Topic", "From Date", "To Date", "Model",
                         "Summary", "Highlights", "Consensus"])

    ws_news_sources = wb.create_sheet("News Sources")
    ws_news_sources.append(["Timestamp", "Topic", "Source"])

    ws_news_quotes = wb.create_sheet("News Quotes")
    ws_news_quotes.append(["Timestamp", "Topic", "Quote", "Tag"])

# ── Runs sheet: one row per run ────────────────────────────────────────────────

ws_runs.append([
    timestamp,
    TOPIC,
    FROM_DATE.strftime("%Y-%m-%d"),
    TO_DATE.strftime("%Y-%m-%d"),
    MODEL,
    x_final.summary,
    x_final.highlights,
    x_final.consensus,
])

# ── Quotes sheet: one row per quote, exact duplicates excluded ─────────────────

def normalize(s: str) -> str:
    """Standardize quote characters for comparison only — don't alter stored text."""
    return s.replace('"', "'")

existing_quotes = {
    normalize(ws_quotes.cell(row=r, column=3).value)
    for r in range(2, ws_quotes.max_row + 1)
}

added_quotes = 0
for quote in x_final.quotes:
    if normalize(quote.text) not in existing_quotes:
        ws_quotes.append([timestamp, TOPIC, quote.text, quote.tag])
        existing_quotes.add(normalize(quote.text))
        added_quotes += 1

print(f"\nX quotes: {added_quotes} new added ({len(x_final.quotes) - added_quotes} duplicates skipped)")

# ── News Runs sheet: one row per run ──────────────────────────────────────────

ws_news_runs.append([
    timestamp,
    TOPIC,
    FROM_DATE.strftime("%Y-%m-%d"),
    TO_DATE.strftime("%Y-%m-%d"),
    MODEL,
    news_final.summary,
    news_final.highlights,
    news_final.consensus,
])

# ── News Sources sheet: one row per source, exact duplicates excluded ──────────

existing_sources = {
    normalize(ws_news_sources.cell(row=r, column=3).value)
    for r in range(2, ws_news_sources.max_row + 1)
}

new_sources = [s.strip() for s in news_final.sources.splitlines() if s.strip()]

added_sources = 0
for source in new_sources:
    if normalize(source) not in existing_sources:
        ws_news_sources.append([timestamp, TOPIC, source])
        existing_sources.add(normalize(source))
        added_sources += 1

# ── News Quotes sheet: one row per quote, exact duplicates excluded ────────────

existing_news_quotes = {
    normalize(ws_news_quotes.cell(row=r, column=3).value)
    for r in range(2, ws_news_quotes.max_row + 1)
}

added_news_quotes = 0
for nq in news_final.news_quotes:
    if normalize(nq.text) not in existing_news_quotes:
        ws_news_quotes.append([timestamp, TOPIC, nq.text, nq.tag])
        existing_news_quotes.add(normalize(nq.text))
        added_news_quotes += 1

wb.save(EXCEL_PATH)
print(f"News sources: {added_sources} new added ({len(new_sources) - added_sources} duplicates skipped)")
print(f"News quotes: {added_news_quotes} new added ({len(news_final.news_quotes) - added_news_quotes} duplicates skipped)")
print(f"Saved to {EXCEL_PATH}")
